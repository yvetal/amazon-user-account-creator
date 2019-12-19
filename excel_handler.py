import openpyxl
print('Excel Handler')
def read_excel(path):
	print('Reading Excel From Path '+path)
	wb=openpyxl.load_workbook(path)
	sheet=wb.active
	row0=list(sheet.rows)[0]
	n=len(row0)
	details=[]
	for row in list(sheet.rows)[1:]:
		j=0
		map={}
		for cell in row:
			if j>=n:
				break
			elif row0[j].value!=None and cell.value!=None:
				map[row0[j].value]=cell.value
			j=j+1
		if map:
			details.append(map)
	return details	

def read_created_accounts():
	return read_excel('credential_details/created_accounts.xlsx')

def read_amazon_credentials():
	return read_excel('credential_details/amazon_credentials.xlsx')
	
def read_email_credentials():
	return read_excel('credential_details/email_credentials.xlsx')

def write_excel(path, contents):
	columns=list(contents[0].keys())
	wb = openpyxl.Workbook()
	sheet=wb.active
	
	for i in range(1,len(columns)+1):
		sheet.cell(row=1, column=i).value=columns[i-1]
	i=2
	for entry in contents:
		for j in range(1,len(columns)+1):
			sheet.cell(row=i, column=j).value=entry[columns[j-1]]
		i=i+1
	wb.save(path)
	
def write_created_accounts(created_accounts_details):
	write_excel('credential_details/created_accounts.xlsx',created_accounts_details)